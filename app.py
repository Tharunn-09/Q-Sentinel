# ==============================================================================
# Q-Sentinel Backend API
# Run: python app.py
# ==============================================================================

from flask import Flask, request, jsonify, session
from flask_cors import CORS
import socket
import ssl
import datetime
import random
import string
import hashlib
import json
import pyotp
import urllib.request
import urllib.parse
from cryptography import x509
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives.asymmetric import rsa, ec, dsa
import groq
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

import database_mongo as db

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "qsentinel-mfa-secret-2026")


# ┌────────────────────────────────────────────────────────┐
# │  API KEYS                                              │
# └────────────────────────────────────────────────────────┘

# ADD YOUR GROQ API KEY HERE (line below)
# Get your API key from: https://console.groq.com/
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")

# Configure Groq client
client = groq.Groq(api_key=GROQ_API_KEY)
API_KEY = "qsentinel-team-vectorx-2026"


def add_options_route(route_func):
    """Wrap a route to also handle OPTIONS preflight requests."""
    original = route_func
    def wrapper(*args, **kwargs):
        if request.method == "OPTIONS":
            response = app.make_response("")
            response.headers["Access-Control-Allow-Origin"] = "*"
            response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
            response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-API-Key"
            return response
        return original(*args, **kwargs)
    wrapper.__name__ = original.__name__
    return wrapper

# Apply OPTIONS handler to all API routes via before_request
@app.before_request
def handle_options():
    if request.method == "OPTIONS":
        response = app.make_response("")
        response.headers["Access-Control-Allow-Origin"] = "*"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-API-Key"
        return response

# Allow frontend HTML to communicate with this backend
CORS(app, resources={r"/*": {"origins": "*"}})

def check_api_key():
    """Validate the frontend API key."""
    return request.headers.get("X-API-Key", "") == API_KEY

# ─ 1. Serve Frontend & Health Endpoints ──────────────────────────────────────────
@app.route("/", methods=["GET"])
def index():
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        html_path = os.path.join(base_dir, "Q-Sentinel_PNB.html")
        with open(html_path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        return f"Error loading frontend: {str(e)}", 500

@app.route("/pnb_logo.jpg", methods=["GET"])
def serve_pnb_logo():
    from flask import send_from_directory
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return send_from_directory(base_dir, "pnb_logo.jpg")

@app.route("/logo.png", methods=["GET"])
def serve_logo():
    from flask import send_from_directory
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return send_from_directory(base_dir, "logo.png")

@app.route("/uco_logo.png", methods=["GET"])
def serve_uco_logo():
    from flask import send_from_directory
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return send_from_directory(base_dir, "uco_logo.png")

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "service": "Q-Sentinel Backend Online"})

# ─ 2. TLS Scanner Endpoint ────────────────────────────────────────────────────
@app.route("/scan", methods=["POST"])
def scan():
    print(f"\n[SCAN] Received scan request")
    if not check_api_key():
        print(f"[SCAN] Unauthorized - invalid API key")
        return jsonify({"error": "Unauthorized — invalid or missing API key"}), 401

    data = request.get_json(force=True, silent=True) or {}
    hostname = (data.get("hostname") or "").strip()
    port     = int(data.get("port", 443))

    print(f"[SCAN] Scanning {hostname}:{port}")
    if not hostname:
        print(f"[SCAN] Missing hostname")
        return jsonify({"error": "Hostname is required"}), 400

    try:
        result = analyze_tls(hostname, port)

        # Print scan details to terminal (VS Code terminal)
        print("\n" + "="*60)
        print(f"  TLS SCAN RESULTS FOR: {hostname}")
        print("="*60)
        print(f"  IP Address    : {result.get('ipAddress', 'N/A')}")
        print(f"  TLS Version   : {result.get('tls', 'N/A')}")
        print(f"  Cipher        : {result.get('cipher', 'N/A')}")
        print(f"  Cert Algo     : {result.get('certAlgo', 'N/A')} ({result.get('keySize', 'N/A')}-bit)")
        print(f"  Issuer        : {result.get('issuer', 'N/A')}")
        print(f"  Subject       : {result.get('subject', 'N/A')}")
        print(f"  Serial Number : {result.get('serialNumber', 'N/A')}")
        print(f"  Valid From    : {result.get('validFrom', 'N/A')}")
        print(f"  Expiry        : {result.get('expiry', 'N/A')}")
        print(f"  Days Left     : {result.get('daysLeft', 'N/A')}")
        print(f"  Key Strength  : {result.get('keyStrength', 'N/A')}")
        print(f"  Key Exchange  : {result.get('keyExchange', 'N/A')}")
        print(f"  Fingerprint   : {result.get('fingerprint', 'N/A')}")
        print(f"  Signature Algo: {result.get('signatureAlgorithm', 'N/A')}")
        print(f"  SSL Grade     : {result.get('sslGrade', 'N/A')}")
        print(f"  Quantum Risk  : {result.get('quantumRisk', 'N/A')}")
        print("="*60 + "\n")

        # Store scan results in database
        try:
            scan_id, host_id = db.store_scan_results(result, hostname, port)
            result['scan_id'] = str(scan_id)
            result['host_id'] = str(host_id)
            result['db_stored'] = True
        except Exception as db_error:
            result['db_stored'] = False
            result['db_error'] = str(db_error)

        return jsonify(result)
    except Exception as e:
        print(f"[SCAN] Error scanning {hostname}: {str(e)}")
        return jsonify({"error": str(e)}), 500

def analyze_tls(hostname, port):
    """Performs live TLS handshake and extracts cert details."""
    context = ssl.create_default_context()
    context.check_hostname = False
    context.verify_mode = ssl.CERT_NONE

    with socket.create_connection((hostname, port), timeout=5) as sock:
        with context.wrap_socket(sock, server_hostname=hostname) as ssock:
            cipher_info = ssock.cipher()
            cipher_name = cipher_info[0]
            tls_version = cipher_info[1].replace('v', ' ')
            der_cert = ssock.getpeercert(binary_form=True)

    cert = x509.load_der_x509_certificate(der_cert, default_backend())

    issuer = "Unknown CA"
    issuer_attrs_cn = cert.issuer.get_attributes_for_oid(x509.NameOID.COMMON_NAME)
    if issuer_attrs_cn:
        issuer = issuer_attrs_cn[0].value

    try:
        not_after = cert.not_valid_after_utc
    except AttributeError:
        not_after = cert.not_valid_after.replace(tzinfo=datetime.timezone.utc)
        
    now = datetime.datetime.now(datetime.timezone.utc)
    days_left = (not_after - now).days
    months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    expiry_str = f"{not_after.day:02d} {months[not_after.month-1]} {not_after.year}"

    public_key = cert.public_key()
    if isinstance(public_key, rsa.RSAPublicKey):
        cert_algo = "RSA"
        key_size = public_key.key_size
    elif isinstance(public_key, ec.EllipticCurvePublicKey):
        cert_algo = "ECDSA"
        key_size = public_key.curve.key_size
    else:
        cert_algo = "UNKNOWN"
        key_size = 0

    quantum_risk = "LOW"
    if tls_version in ["TLS 1.0", "TLS 1.1", "SSLv3"] or key_size <= 1024 or "RC4" in cipher_name:
        quantum_risk = "CRITICAL"
    elif "CBC" in cipher_name or (cert_algo == "RSA" and key_size <= 2048):
        quantum_risk = "HIGH"
    elif cert_algo == "RSA" and key_size <= 3072:
        quantum_risk = "MODERATE"

    ssl_grade = "B"
    if days_left < 0 or tls_version in ["TLS 1.0", "SSLv3"] or key_size < 1024:
        ssl_grade = "F"
    elif tls_version == "TLS 1.1":
        ssl_grade = "C"
    elif tls_version == "TLS 1.3" and key_size >= 2048 and days_left > 30:
        ssl_grade = "A+"

    # Get additional certificate details
    subject = cert.subject.rfc4514_string()
    serial_number = str(cert.serial_number)

    try:
        sig_algo = cert.signature_hash_algorithm.name + "With" + type(cert.public_key()).__name__
    except Exception:
        sig_algo = str(cert.signature_algorithm_oid)

    # Get validity dates
    try:
        not_before = cert.not_valid_before_utc
    except AttributeError:
        not_before = cert.not_valid_before.replace(tzinfo=datetime.timezone.utc)

    months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    valid_from_str = f"{not_before.day:02d} {months[not_before.month-1]} {not_before.year}"

    # Calculate key strength
    key_strength = "UNKNOWN"
    if cert_algo == "RSA":
        if key_size < 2048:
            key_strength = "WEAK"
        elif key_size < 3072:
            key_strength = "ACCEPTABLE"
        else:
            key_strength = "STRONG"
    elif cert_algo == "ECDSA":
        if key_size < 256:
            key_strength = "WEAK"
        elif key_size < 384:
            key_strength = "ACCEPTABLE"
        else:
            key_strength = "STRONG"

    # Get IP address
    try:
        ip_address = socket.gethostbyname(hostname)
    except Exception:
        ip_address = "Unknown"

    # Generate fingerprint
    fingerprint = ":".join("{:02x}".format(b) for b in der_cert[:20])

    # Key exchange method
    key_exchange = "Unknown"
    cipher_upper = cipher_name.upper()
    if "ECDHE" in cipher_upper:
        key_exchange = "ECDHE (Elliptic Curve Diffie-Hellman Ephemeral)"
    elif "DHE" in cipher_upper or "EDH" in cipher_upper:
        key_exchange = "DHE (Diffie-Hellman Ephemeral)"
    elif "TLS_AES" in cipher_upper or "TLS_CHACHA" in cipher_upper:
        key_exchange = "X25519 / P-256 (TLS 1.3 Ephemeral)"
    elif "RSA" in cipher_upper:
        key_exchange = "RSA (Static - No Forward Secrecy)"
    elif tls_version == "TLS 1.3":
        key_exchange = "X25519 / P-256 (TLS 1.3 Ephemeral)"

    return {
        "host": hostname,
        "port": port,
        "ipAddress": ip_address,
        "tls": tls_version,
        "cipher": cipher_name,
        "certAlgo": cert_algo,
        "keySize": key_size,
        "issuer": issuer,
        "expiry": expiry_str,
        "daysLeft": days_left,
        "quantumRisk": quantum_risk,
        "sslGrade": ssl_grade,
        "pqcDetected": False,
        "errors": [],
        "subject": subject,
        "serialNumber": serial_number,
        "signatureAlgorithm": sig_algo,
        "validFrom": valid_from_str,
        "keyStrength": key_strength,
        "fingerprint": fingerprint,
        "keyExchange": key_exchange
    }

def evaluate_subdomain_tls(hostname, port=443):
    """Helper to perform a quick, light TLS risk assessment on a subdomain."""
    try:
        # Resolve IP address
        try:
            ip_address = socket.gethostbyname(hostname)
        except Exception:
            ip_address = "Unknown"

        context = ssl.create_default_context()
        context.check_hostname = False
        context.verify_mode = ssl.CERT_NONE

        # Short timeout for fast subdomain scanning
        with socket.create_connection((hostname, port), timeout=2.5) as sock:
            with context.wrap_socket(sock, server_hostname=hostname) as ssock:
                cipher_info = ssock.cipher()
                cipher_name = cipher_info[0]
                tls_version = cipher_info[1].replace('v', ' ')
                der_cert = ssock.getpeercert(binary_form=True)

        cert = x509.load_der_x509_certificate(der_cert, default_backend())

        public_key = cert.public_key()
        if isinstance(public_key, rsa.RSAPublicKey):
            cert_algo = "RSA"
            key_size = public_key.key_size
        elif isinstance(public_key, ec.EllipticCurvePublicKey):
            cert_algo = "ECDSA"
            key_size = public_key.curve.key_size
        else:
            cert_algo = "UNKNOWN"
            key_size = 0

        # Calculate certificate expiry
        try:
            not_after = cert.not_valid_after_utc
        except AttributeError:
            not_after = cert.not_valid_after.replace(tzinfo=datetime.timezone.utc)
        now = datetime.datetime.now(datetime.timezone.utc)
        days_left = (not_after - now).days

        quantum_risk = "LOW"
        if tls_version in ["TLS 1.0", "TLS 1.1", "SSLv3"] or key_size <= 1024 or "RC4" in cipher_name:
            quantum_risk = "CRITICAL"
        elif "CBC" in cipher_name or (cert_algo == "RSA" and key_size <= 2048):
            quantum_risk = "HIGH"
        elif cert_algo == "RSA" and key_size <= 3072:
            quantum_risk = "MODERATE"

        ssl_grade = "B"
        if days_left < 0 or tls_version in ["TLS 1.0", "SSLv3"] or key_size < 1024:
            ssl_grade = "F"
        elif tls_version == "TLS 1.1":
            ssl_grade = "C"
        elif tls_version == "TLS 1.3" and key_size >= 2048 and days_left > 30:
            ssl_grade = "A+"

        details = f"{tls_version}, {cert_algo} {key_size}-bit"
        return {
            "name": hostname,
            "ipAddress": ip_address,
            "sslGrade": ssl_grade,
            "quantumRisk": quantum_risk,
            "details": details
        }
    except Exception as e:
        try:
            ip_address = socket.gethostbyname(hostname)
        except Exception:
            ip_address = "Unknown"
        return {
            "name": hostname,
            "ipAddress": ip_address,
            "sslGrade": "N/A",
            "quantumRisk": "UNKNOWN",
            "details": "Offline or port 443 closed"
        }

# ─ 3. Subfinder Subdomain Discovery ──────────────────────────────────────────
@app.route("/api/subfinder-subdomains", methods=["GET"])
def subfinder_subdomains():
    """Use Subfinder (passive) to discover subdomains for a given hostname."""
    hostname = request.args.get("hostname", "").strip()
    if not hostname:
        return jsonify({"error": "hostname is required"}), 400

    # Extract base domain
    parts = hostname.lower().strip().split(".")
    if len(parts) < 2:
        return jsonify({"error": "invalid hostname"}), 400

    two_part_tlds = ["co.in", "org.in", "net.in", "bank.in", "gov.in", "edu.in", "co.uk", "org.uk", "com.au", "ac.in", "nic.in"]
    joined2 = ".".join(parts[-2:])
    joined3 = ".".join(parts[-3:]) if len(parts) >= 3 else ""

    if len(parts) >= 3 and any(tld in joined2 or joined3.endswith("." + tld) for tld in two_part_tlds):
        base_domain = ".".join(parts[-3:])
    else:
        base_domain = ".".join(parts[-2:])

    print(f"[SUBFINDER] Enumerating subdomains for: {base_domain}")

    subdomains_set = set()
    
    # 1. Try Subfinder
    try:
        import subprocess, json as json_mod
        result = subprocess.run(
            [r"C:\tools\subfinder\subfinder.exe", "-d", base_domain, "-silent", "-json"],
            capture_output=True, text=True, timeout=60, shell=False
        )
        if result.returncode == 0 and result.stdout.strip():
            for line in result.stdout.strip().split("\n"):
                try:
                    entry = json_mod.loads(line)
                    sub = (entry.get("host") or entry.get("name") or "").strip().lower()
                    if sub and sub.endswith("." + base_domain):
                        subdomains_set.add(sub)
                except Exception:
                    pass
        elif result.stderr:
            print(f"[SUBFINDER] stderr: {result.stderr[:200]}")
    except subprocess.TimeoutExpired:
        print("[SUBFINDER] Timed out after 60s")
    except Exception as e:
        print(f"[SUBFINDER] Subfinder binary failed or not found: {str(e)}. Using OSINT sources.")

    # 2. Try crt.sh (Certificate Transparency Logs)
    try:
        import urllib.request, json as json_mod
        url = f"https://crt.sh/?q=%.{base_domain}&output=json"
        req = urllib.request.Request(
            url, 
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        )
        with urllib.request.urlopen(req, timeout=15) as response:
            data = json_mod.loads(response.read().decode('utf-8'))
            for entry in data:
                name_value = entry.get("name_value", "")
                for sub in name_value.split():
                    sub = sub.strip().lower()
                    if sub.startswith("*."):
                        sub = sub[2:]
                    if sub.endswith("." + base_domain):
                        subdomains_set.add(sub)
    except Exception as e:
        print(f"[SUBFINDER-OSINT] crt.sh failed: {str(e)}")

    # 3. Try HackerTarget passive DNS
    try:
        import urllib.request
        url = f"https://api.hackertarget.com/hostsearch/?q={base_domain}"
        req = urllib.request.Request(
            url, 
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req, timeout=10) as response:
            content = response.read().decode('utf-8')
            for line in content.split("\n"):
                parts = line.split(",")
                if len(parts) >= 1:
                    sub = parts[0].strip().lower()
                    if sub.endswith("." + base_domain):
                        subdomains_set.add(sub)
    except Exception as e:
        print(f"[SUBFINDER-OSINT] HackerTarget failed: {str(e)}")

    # 4. Try Anubis DB
    try:
        import urllib.request, json as json_mod
        url = f"https://jldc.me/anubis/subdomains/{base_domain}"
        req = urllib.request.Request(
            url, 
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json_mod.loads(response.read().decode('utf-8'))
            if isinstance(data, list):
                for sub in data:
                    sub = sub.strip().lower()
                    if sub.endswith("." + base_domain):
                        subdomains_set.add(sub)
    except Exception as e:
        print(f"[SUBFINDER-OSINT] Anubis failed: {str(e)}")

    # 5. Try AlienVault OTX
    try:
        import urllib.request, json as json_mod
        url = f"https://otx.alienvault.com/api/v1/indicators/domain/{base_domain}/passive_dns"
        req = urllib.request.Request(
            url, 
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json_mod.loads(response.read().decode('utf-8'))
            records = data.get("passive_dns", [])
            for record in records:
                sub = record.get("hostname", "").strip().lower()
                if sub.endswith("." + base_domain):
                    subdomains_set.add(sub)
    except Exception as e:
        print(f"[SUBFINDER-OSINT] AlienVault OTX failed: {str(e)}")

    subdomains = sorted(list(subdomains_set))
    print(f"[SUBFINDER-COMBINED] Found a total of {len(subdomains)} unique subdomains for {base_domain}")
    print(f"[SUBFINDER] Evaluating risks for {len(subdomains)} subdomains...")
    
    evaluated_subdomains = []
    if subdomains:
        import concurrent.futures
        with concurrent.futures.ThreadPoolExecutor(max_workers=15) as executor:
            results = executor.map(evaluate_subdomain_tls, subdomains)
            evaluated_subdomains = list(results)

    print(f"[SUBFINDER] Finished evaluating risks for {len(subdomains)} subdomains.")
    return jsonify({
        "hostname": hostname,
        "base_domain": base_domain,
        "subdomains": evaluated_subdomains,
        "count": len(evaluated_subdomains)
    })

# ─ 4. Store Scan (Fallback Storage) ──────────────────────────────────────────
@app.route("/api/store-scan", methods=["POST"])
def store_scan():
    """Store scan results from frontend fallback paths (SSL Labs, crt.sh, default)."""
    if not check_api_key():
        return jsonify({"error": "Unauthorized — invalid or missing API key"}), 401

    data = request.get_json(force=True, silent=True) or {}
    hostname = (data.get("host") or data.get("hostname") or "").strip()
    port = int(data.get("port", 443))

    if not hostname:
        return jsonify({"error": "hostname is required"}), 400

    try:
        # Normalize field names to match what store_scan_results expects
        scan_payload = {
            "host":               hostname,
            "port":               port,
            "ipAddress":          data.get("ipAddress", ""),
            "tls":                data.get("tls", "TLS 1.2"),
            "cipher":             data.get("cipher", ""),
            "certAlgo":           data.get("certAlgo", "RSA"),
            "keySize":            data.get("keySize", 2048),
            "issuer":             data.get("issuer", "Unknown"),
            "expiry":             data.get("expiry", "Unknown"),
            "validFrom":          data.get("validFrom", "Unknown"),
            "daysLeft":           data.get("daysLeft", -1),
            "quantumRisk":        data.get("quantumRisk", "UNKNOWN"),
            "sslGrade":           data.get("sslGrade", "?"),
            "keyStrength":        data.get("keyStrength", "UNKNOWN"),
            "keyExchange":        data.get("keyExchange", ""),
            "subject":            data.get("subject", ""),
            "serialNumber":       data.get("serialNumber", ""),
            "signatureAlgorithm": data.get("signatureAlgorithm", ""),
            "fingerprint":        data.get("fingerprint", ""),
            "pqcDetected":        data.get("pqcDetected", False),
            "errors":             data.get("errors", []),
            "scan_source":        data.get("scan_source", "frontend-fallback"),
        }

        print(f"\n[STORE-SCAN] Storing fallback scan for {hostname}:{port} (source: {scan_payload['scan_source']})")

        scan_id, host_id = db.store_scan_results(scan_payload, hostname, port)

        print(f"[STORE-SCAN] ✓ Stored — scan_id={scan_id}, host_id={host_id}")

        return jsonify({
            "status":     "stored",
            "scan_id":    str(scan_id),
            "host_id":    str(host_id),
            "hostname":   hostname,
            "db_stored":  True
        })

    except Exception as e:
        print(f"[STORE-SCAN] Error: {e}")
        return jsonify({"error": str(e), "db_stored": False}), 500


# ─ 4. Groq AI Assistant Endpoint ─────────────────────────────────────────────
@app.route("/api/assistant", methods=["POST"])
def ai_assistant():
    """Handles chat requests from the frontend and sends them to Groq."""
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401

    data = request.get_json(force=True, silent=True) or {}
    question = data.get("question", "").strip()
    scan_data = data.get("scan_data", {})

    if not question:
        return jsonify({"error": "Please provide a question."}), 400

    try:
        # 1. Give Groq the context of the scan
        system_context = f"""You are a Post-Quantum Cryptography expert helping a user secure their web assets.
You are currently analyzing this asset: {scan_data.get('host', 'No asset scanned yet')}
Risk Level: {scan_data.get('quantumRisk', 'Unknown')}
Algorithm: {scan_data.get('certAlgo', 'Unknown')} ({scan_data.get('keySize', 'Unknown')}-bit)

Grading Rubric / Criteria Details for Q-Sentinel:
1. Quantum Risk Grading:
   - CRITICAL: Legacy TLS (TLS 1.0, 1.1, SSLv3), weak key size (<= 1024-bit), or weak ciphers (RC4).
   - HIGH: CBC mode ciphers (vulnerable to padding oracle attacks) or RSA key size <= 2048-bit.
   - MODERATE: RSA key size <= 3072-bit.
   - LOW: Modern secure cryptography (e.g. ECDSA curves, or RSA >= 4096-bit).
2. SSL Grade Grading:
   - F: Certificate expired (days left < 0), legacy TLS (TLS 1.0, SSLv3), or key size < 1024-bit.
   - C: TLS 1.1 version.
   - B: Default/Acceptable (TLS 1.2, standard configurations).
   - A+: TLS 1.3, key size >= 2048-bit, and certificate valid for > 30 days.

Please explain the scanned assets, discovered domains/subdomains, the risks associated with them, their counts, and why they received their specific grades (SSL Grade or Quantum Risk) based on the above criteria in a clear, technical, and professional tone."""

        # 2. Call the Groq API using chat completions
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": system_context},
                {"role": "user", "content": question}
            ],
            temperature=0.7,
            max_tokens=1024
        )

        return jsonify({
            "status": "success",
            "answer": response.choices[0].message.content
        })

    except Exception as e:
        return jsonify({"error": f"Groq API Error: {str(e)}"}), 500

# Initialize database
db.init_database()
print(f"[DB] Q-Sentinel MongoDB database initialized")

# ─ 4. Get Last Scan for Host ─────────────────────────────────────────────────
@app.route("/api/last-scan/<hostname>", methods=["GET"])
def get_last_scan(hostname):
    """Get the most recent scan results for a host."""
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401

    try:
        result = db.get_last_scan_for_host(hostname)
        if result:
            return jsonify(result)
        return jsonify({"error": "No scans found for host"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ─ 5. Get Scan History ────────────────────────────────────────────────────────
@app.route("/api/history/<hostname>", methods=["GET"])
def get_history(hostname):
    """Get scan history for a host."""
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401

    limit = request.args.get('limit', 10, type=int)
    try:
        history = db.get_scan_history(hostname, limit)
        return jsonify({"hostname": hostname, "history": history})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ─ 6. Get Comprehensive Report ────────────────────────────────────────────────
@app.route("/api/report/<hostname>", methods=["GET"])
def get_report(hostname):
    """Get comprehensive CBOM, PQC risk, and scan report for a host."""
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401

    try:
        report = db.get_comprehensive_report(hostname)
        if report:
            return jsonify(report)
        return jsonify({"error": "No report found for host"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ─ 7. Get Dashboard Summary ───────────────────────────────────────────────────
@app.route("/api/dashboard", methods=["GET"])
def get_dashboard():
    """Get dashboard summary statistics."""
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401

    try:
        summary = db.get_dashboard_summary()
        return jsonify(summary)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ─ 8. Get All Hosts ───────────────────────────────────────────────────────────
@app.route("/api/hosts", methods=["GET"])
def get_hosts():
    """Get all registered hosts."""
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401

    try:
        hosts = db.get_all_hosts()
        return jsonify({"hosts": hosts})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ─ 9. Export CBOM ─────────────────────────────────────────────────────────────
@app.route("/api/export/cbom/<hostname>", methods=["GET"])
def export_cbom(hostname):
    """Export CBOM as JSON."""
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401

    try:
        cbom = db.export_cbom_json(hostname)
        if cbom:
            return jsonify(cbom)
        return jsonify({"error": "No CBOM found for host"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500





# ═══════════════════════════════════════════════════════════════
#  USER AUTHENTICATION & MFA ENDPOINTS (MongoDB-Backed)
# ═══════════════════════════════════════════════════════════════

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Load secrets from environment (or VoxShield fallback)
JWT_SECRET = os.environ.get("JWT_SECRET", "voxshield-secure-token-secret-key-2026-xyz")
GMAIL_USER = os.environ.get("GMAIL_USER", "voxshield.auth@gmail.com")
GMAIL_APP_PASS = os.environ.get("GMAIL_APP_PASS", "iakzmohjtwhwxsru")
RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")

def send_smtp_email(to_email, subject, text_content, html_content, attachment_data=None, attachment_filename=None):
    """Utility to send an email using Gmail SMTP, optionally with an attachment."""
    try:
        from email.mime.base import MIMEBase
        from email import encoders
        if attachment_data:
            msg = MIMEMultipart('mixed')
            msg['From'] = f'"Q-Sentinel Security" <{GMAIL_USER}>'
            msg['To'] = to_email
            msg['Subject'] = subject

            alt_part = MIMEMultipart('alternative')
            part1 = MIMEText(text_content, 'plain')
            part2 = MIMEText(html_content, 'html')
            alt_part.attach(part1)
            alt_part.attach(part2)
            msg.attach(alt_part)

            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment_data)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{attachment_filename or "report.pdf"}"')
            msg.attach(part)
        else:
            msg = MIMEMultipart('alternative')
            msg['From'] = f'"Q-Sentinel Security" <{GMAIL_USER}>'
            msg['To'] = to_email
            msg['Subject'] = subject

            part1 = MIMEText(text_content, 'plain')
            part2 = MIMEText(html_content, 'html')
            msg.attach(part1)
            msg.attach(part2)

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(GMAIL_USER, GMAIL_APP_PASS)
        server.sendmail(GMAIL_USER, to_email, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        print(f"[SMTP ERROR] Failed to send email: {e}")
        return False

def send_email(to_email, subject, text_content, html_content, attachment_data=None, attachment_filename=None):
    """Send an email using Resend API if key is present, otherwise fallback to SMTP."""
    if RESEND_API_KEY:
        try:
            import urllib.request
            import json
            import base64
            url = "https://api.resend.com/emails"
            headers = {
                "Authorization": f"Bearer {RESEND_API_KEY}",
                "Content-Type": "application/json"
            }
            data = {
                "from": "Q-Sentinel Security <onboarding@resend.dev>",
                "to": [to_email],
                "subject": subject,
                "text": text_content,
                "html": html_content
            }
            if attachment_data:
                b64_content = base64.b64encode(attachment_data).decode('utf-8')
                data["attachments"] = [{
                    "content": b64_content,
                    "filename": attachment_filename or "report.pdf"
                }]
            req = urllib.request.Request(
                url,
                data=json.dumps(data).encode("utf-8"),
                headers=headers,
                method="POST"
            )
            with urllib.request.urlopen(req) as response:
                res_data = json.loads(response.read().decode())
                if response.status == 200 or res_data.get("id"):
                    print(f"[EMAIL] Successfully sent email to {to_email} via Resend API.")
                    return True
        except Exception as e:
            print(f"[RESEND ERROR] Failed to send email via Resend, falling back to SMTP: {e}")
            
    return send_smtp_email(to_email, subject, text_content, html_content, attachment_data, attachment_filename)

def generate_signed_token(payload, expires_in_seconds=300):
    """Generate a simple signed token resembling a JWT."""
    import time
    import hmac
    payload_copy = payload.copy()
    payload_copy["exp"] = int(time.time()) + expires_in_seconds
    payload_str = json.dumps(payload_copy, sort_keys=True)
    signature = hmac.new(JWT_SECRET.encode(), payload_str.encode(), hashlib.sha256).hexdigest()
    return f"{payload_str}.{signature}"

def verify_signed_token(token):
    """Verify and decode a signed token."""
    import time
    import hmac
    try:
        parts = token.split(".")
        if len(parts) != 2:
            return None
        payload_str, signature = parts[0], parts[1]
        expected_signature = hmac.new(JWT_SECRET.encode(), payload_str.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected_signature):
            return None
        payload = json.loads(payload_str)
        if int(time.time()) > payload.get("exp", 0):
            return None
        return payload
    except Exception:
        return None

def hash_password(password):
    """Hash password using SHA-256 with salt."""
    salt = "qsentinel_salt_2026_"
    return hashlib.sha256((salt + password).encode()).hexdigest()

def seed_default_user():
    """Seed default Q-Sentinel Agent if not exists."""
    try:
        existing = db.get_user("PNB-AGT-1042")
        if not existing:
            hashed = hash_password("pnbbank@2026")
            user_doc = {
                "employeeId": "PNB-AGT-1042",
                "name": "PNB Agent 1042",
                "email": "agent1042@pnb.co.in",
                "password": hashed,
                "mfaSecret": "ABCDEFGHIJKLMNOPQRST", # 20 character base32
                "mfaEnrolled": False,
                "status": "active",
                "otp": {"code": "", "expiresAt": None},
                "loginAttempts": {"count": 0, "lockUntil": None}
            }
            db.save_user(user_doc)
            print("[AUTH] Seeded default agent PNB-AGT-1042 successfully.")
    except Exception as e:
        print(f"[AUTH ERROR] Failed to seed default user: {e}")

# Seed user on init
db.init_database()
seed_default_user()

def is_strong_password(password):
    if len(password) < 8: return False
    has_upper = any(c.isupper() for c in password)
    has_lower = any(c.islower() for c in password)
    has_digit = any(c.isdigit() for c in password)
    has_special = any(not c.isalnum() for c in password)
    return has_upper and has_lower and has_digit and has_special

@app.route("/api/auth/register-init", methods=["POST"])
def register_init():
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401

    data = request.get_json(force=True, silent=True) or {}
    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip()
    user = (data.get("user") or "").strip()
    pass_val = data.get("pass", "")

    if not name or not email or not user or not pass_val:
        return jsonify({"error": "All fields are required"}), 400

    if not is_strong_password(pass_val):
        return jsonify({"error": "Password must be at least 8 characters long and contain 1 uppercase, 1 lowercase, 1 number, and 1 special character."}), 400

    try:
        existing_user = db.get_user(user)
        if existing_user and existing_user.get("status") == "active":
            return jsonify({"error": "Employee ID is already registered"}), 400

        existing_email = db.get_user_by_email(email)
        if existing_email and existing_email.get("status") == "active":
            return jsonify({"error": "Email address is already registered"}), 400

        hashed_password = hash_password(pass_val)
        otp_code = str(random.randint(100000, 999999))
        otp_expires = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(minutes=5)

        if existing_user:
            existing_user["name"] = name
            existing_user["email"] = email
            existing_user["password"] = hashed_password
            existing_user["otp"] = {"code": otp_code, "expiresAt": otp_expires}
            db.save_user(existing_user)
        else:
            new_user = {
                "employeeId": user,
                "name": name,
                "email": email,
                "password": hashed_password,
                "mfaSecret": "",
                "mfaEnrolled": False,
                "status": "pending",
                "otp": {"code": otp_code, "expiresAt": otp_expires},
                "loginAttempts": {"count": 0, "lockUntil": None}
            }
            db.save_user(new_user)

        # Send Email
        subject = "Q-Sentinel Verification Code"
        text_content = f"Hello Officer,\n\nWelcome to Q-Sentinel!\n\nPlease verify your registration with this OTP: {otp_code}\n\nValid for 5 minutes."
        html_content = f"""
            <div style="font-family: Arial, sans-serif; max-width: 500px; margin: 0 auto; padding: 20px; border: 1px solid #e0e0e0; border-radius: 8px; background-color: #0d0a0b; color: #f5e6d3;">
              <h2 style="color: #e8b84b; text-align: center;">Q-Sentinel Security</h2>
              <hr style="border: none; border-top: 1px solid rgba(232, 184, 75, 0.2); margin: 20px 0;" />
              <p>Hello Officer,</p>
              <p>Please enter the 6-digit One-Time Password (OTP) below to verify your email address and authorize your officer profile creation:</p>
              <div style="background-color: #1a1014; border: 1px solid rgba(232, 184, 75, 0.4); padding: 18px; text-align: center; font-size: 32px; font-weight: bold; letter-spacing: 6px; color: #e8b84b; margin: 24px 0; border-radius: 8px; font-family: monospace;">
                {otp_code}
              </div>
              <p style="font-size: 11px; color: #8a7060; text-align: center;">Punjab National Bank Cybersecurity Division</p>
            </div>
        """
        send_email(email, subject, text_content, html_content)
        db.log_audit_event(user, "REGISTRATION_OTP_SENT", request.remote_addr, f"Verification OTP sent to {email}.")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/auth/register-verify", methods=["POST"])
def register_verify():
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401

    data = request.get_json(force=True, silent=True) or {}
    user = (data.get("user") or "").strip()
    otp = (data.get("otp") or "").strip()

    if not user or not otp:
        return jsonify({"error": "Employee ID and OTP are required"}), 400

    try:
        db_user = db.get_user(user)
        if not db_user or db_user.get("status") != "pending":
            return jsonify({"error": "No pending registration found for this Employee ID"}), 400

        otp_doc = db_user.get("otp", {})
        expires_at = otp_doc.get("expiresAt")
        if expires_at:
            if expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=datetime.timezone.utc)

        if otp_doc.get("code") != otp:
            db.log_audit_event(user, "REGISTRATION_OTP_FAILED", request.remote_addr, "Failed OTP entry during registration.")
            return jsonify({"error": "Incorrect verification code"}), 400

        if expires_at < datetime.datetime.now(datetime.timezone.utc):
            db.log_audit_event(user, "REGISTRATION_OTP_EXPIRED", request.remote_addr, "Expired OTP entry attempt.")
            return jsonify({"error": "OTP has expired"}), 400

        # Generate seed
        alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ234567'
        mfa_secret = ''.join(random.choices(alphabet, k=20))
        
        db_user["status"] = "active"
        db_user["mfaSecret"] = mfa_secret
        db_user["mfaEnrolled"] = False
        db_user["otp"] = {"code": "", "expiresAt": None}
        db.save_user(db_user)

        temp_token = generate_signed_token({"employeeId": db_user["employeeId"], "stage": "mfa"})
        db.log_audit_event(user, "REGISTRATION_COMPLETED", request.remote_addr, "Registration completed. Account activated.")

        return jsonify({"success": True, "mfaSecret": mfa_secret, "tempToken": temp_token})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/auth/login-creds", methods=["POST"])
def login_creds():
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401

    data = request.get_json(force=True, silent=True) or {}
    user = (data.get("user") or "").strip()
    pass_val = data.get("pass", "")

    if not user or not pass_val:
        return jsonify({"error": "Employee ID and password are required"}), 400

    try:
        db_user = db.get_user(user)
        if not db_user or db_user.get("status") != "active":
            db.log_audit_event(user, "LOGIN_FAILED", request.remote_addr, "Failed login attempt with unregistered ID.")
            return jsonify({"error": "Invalid Employee ID or password"}), 401

        attempts = db_user.get("loginAttempts", {"count": 0, "lockUntil": None})
        lock_until = attempts.get("lockUntil")
        if lock_until:
            if lock_until.tzinfo is None:
                lock_until = lock_until.replace(tzinfo=datetime.timezone.utc)
            if lock_until > datetime.datetime.now(datetime.timezone.utc):
                remaining_mins = int((lock_until - datetime.datetime.now(datetime.timezone.utc)).total_seconds() / 60)
                return jsonify({"error": f"Account locked. Try again in {remaining_mins + 1} minute(s)."}), 423

        if hash_password(pass_val) != db_user.get("password"):
            count = attempts.get("count", 0) + 1
            if count >= 5:
                db_user["loginAttempts"] = {
                    "count": count,
                    "lockUntil": datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(minutes=15)
                }
                db.save_user(db_user)
                db.log_audit_event(user, "ACCOUNT_LOCKED", request.remote_addr, "Account automatically locked for 15 minutes.")
                return jsonify({"error": "Invalid credentials. Account locked for 15 minutes."}), 423
            else:
                db_user["loginAttempts"] = {"count": count, "lockUntil": None}
                db.save_user(db_user)
                db.log_audit_event(user, "LOGIN_FAILED", request.remote_addr, f"Failed login credential attempt ({count}/5).")
                return jsonify({"error": f"Invalid credentials. ({5 - count} attempt(s) remaining)"}), 401

        # Success
        db_user["loginAttempts"] = {"count": 0, "lockUntil": None}
        db.save_user(db_user)

        temp_token = generate_signed_token({"employeeId": db_user["employeeId"], "stage": "mfa"})
        db.log_audit_event(user, "CREDS_VERIFICATION_SUCCESS", request.remote_addr, "Credentials verified. Awaiting 2FA.")

        return jsonify({
            "success": True,
            "tempToken": temp_token,
            "mfaEnrolled": db_user.get("mfaEnrolled", False),
            "mfaSecret": db_user.get("mfaSecret") if not db_user.get("mfaEnrolled", False) else None
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/auth/login-mfa", methods=["POST"])
def login_mfa_verify():
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401

    data = request.get_json(force=True, silent=True) or {}
    temp_token = data.get("tempToken", "")
    mfa_code = data.get("mfaCode", "")

    if not temp_token or not mfa_code:
        return jsonify({"error": "Temporary session token and MFA code are required"}), 400

    payload = verify_signed_token(temp_token)
    if not payload or payload.get("stage") != "mfa":
        return jsonify({"error": "Temporary login session expired. Please log in again."}), 401

    try:
        user = payload["employeeId"]
        db_user = db.get_user(user)
        if not db_user or db_user.get("status") != "active":
            return jsonify({"error": "User account not found"}), 401

        # Use pyotp to verify TOTP code
        totp = pyotp.TOTP(db_user["mfaSecret"])
        if not totp.verify(mfa_code, valid_window=1):
            db.log_audit_event(user, "MFA_VERIFICATION_FAILED", request.remote_addr, "Failed 2FA Authenticator token match attempt.")
            return jsonify({"error": "Incorrect code. Check your authenticator app."}), 401

        if not db_user.get("mfaEnrolled", False):
            db_user["mfaEnrolled"] = True
            db.save_user(db_user)
            db.log_audit_event(user, "MFA_ENROLLED", request.remote_addr, "Completed initial TOTP authenticator device setup.")

        final_token = generate_signed_token({"employeeId": db_user["employeeId"], "name": db_user["name"], "email": db_user["email"]}, 43200) # 12h
        db.log_audit_event(user, "LOGIN_SUCCESS", request.remote_addr, "User logged in successfully.")

        return jsonify({"success": True, "token": final_token, "user": {"employeeId": db_user["employeeId"], "name": db_user["name"], "email": db_user["email"]}})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/auth/forgot-password", methods=["POST"])
def forgot_password_request():
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401

    data = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip()

    if not email:
        return jsonify({"error": "Email is required"}), 400

    try:
        db_user = db.get_user_by_email(email)
        if not db_user or db_user.get("status") != "active":
            return jsonify({"error": "Email address not registered"}), 400

        otp_code = str(random.randint(100000, 999999))
        otp_expires = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(minutes=5)
        db_user["otp"] = {"code": otp_code, "expiresAt": otp_expires}
        db.save_user(db_user)

        subject = "Q-Sentinel Password Reset Code"
        text_content = f"Hello Officer,\n\nYou requested a password reset. OTP: {otp_code}\n\nValid for 5 minutes."
        html_content = f"""
            <div style="font-family: Arial, sans-serif; max-width: 500px; margin: 0 auto; padding: 20px; border: 1px solid #e0e0e0; border-radius: 8px; background-color: #0d0a0b; color: #f5e6d3;">
              <h2 style="color: #e8b84b; text-align: center;">Q-Sentinel Security</h2>
              <hr style="border: none; border-top: 1px solid rgba(232, 184, 75, 0.2); margin: 20px 0;" />
              <p>Hello Officer,</p>
              <p>Please enter the 6-digit verification code below to authorize your password reset:</p>
              <div style="background-color: #1a1014; border: 1px solid rgba(232, 184, 75, 0.4); padding: 18px; text-align: center; font-size: 32px; font-weight: bold; letter-spacing: 6px; color: #e8b84b; margin: 24px 0; border-radius: 8px; font-family: monospace;">
                {otp_code}
              </div>
              <p style="font-size: 11px; color: #8a7060; text-align: center;">Punjab National Bank Cybersecurity Division</p>
            </div>
        """
        send_email(email, subject, text_content, html_content)
        db.log_audit_event(db_user["employeeId"], "PASSWORD_RESET_OTP_SENT", request.remote_addr, "Password reset OTP sent.")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/auth/reset-password", methods=["POST"])
def reset_password_confirm():
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401

    data = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip()
    otp = (data.get("otp") or "").strip()
    new_password = data.get("newPassword", "")

    if not email or not otp or not new_password:
        return jsonify({"error": "All fields are required"}), 400

    if not is_strong_password(new_password):
        return jsonify({"error": "Password must be at least 8 characters long and contain 1 uppercase, 1 lowercase, 1 number, and 1 special character."}), 400

    try:
        db_user = db.get_user_by_email(email)
        if not db_user or db_user.get("status") != "active":
            return jsonify({"error": "User not found"}), 404

        otp_doc = db_user.get("otp", {})
        expires_at = otp_doc.get("expiresAt")
        if expires_at:
            if expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=datetime.timezone.utc)

        if otp_doc.get("code") != otp:
            db.log_audit_event(db_user["employeeId"], "PASSWORD_RESET_FAILED", request.remote_addr, "Failed OTP reset entry.")
            return jsonify({"error": "Incorrect verification code"}), 400

        if expires_at < datetime.datetime.now(datetime.timezone.utc):
            db.log_audit_event(db_user["employeeId"], "PASSWORD_RESET_EXPIRED", request.remote_addr, "Expired OTP reset entry attempt.")
            return jsonify({"error": "OTP has expired"}), 400

        new_hashed = hash_password(new_password)
        if db_user.get("password") == new_hashed:
            return jsonify({"error": "New password cannot be the same as the old password"}), 400

        db_user["password"] = new_hashed
        db_user["otp"] = {"code": "", "expiresAt": None}
        db.save_user(db_user)
        db.log_audit_event(db_user["employeeId"], "PASSWORD_RESET_SUCCESS", request.remote_addr, "Password reset successfully completed.")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ═══════════════════════════════════════════════════════════════
#  REPORT GENERATION & SCHEDULER ENGINE
# ═══════════════════════════════════════════════════════════════

import io
import threading
import time

def compute_scan_stats(scan_results):
    if not scan_results:
        return {
            "assets": 0,
            "rating": "--",
            "posture": "0%",
            "cbom": 0
        }
    total = len(scan_results)
    scores = []
    low_count = 0
    cbom_count = 0
    for r in scan_results:
        raw = r.get('raw_response') or {}
        # Replicate client-side exact risk score computation
        tls = r.get('tls') or r.get('tls_version') or raw.get('tls') or raw.get('tls_version') or ''
        cipher = (r.get('cipher') or r.get('cipher_suite') or raw.get('cipher') or raw.get('cipher_suite') or '').upper()
        algo = r.get('certAlgo') or r.get('cert_algo') or raw.get('certAlgo') or raw.get('cert_algo') or r.get('sigAlgo') or raw.get('sigAlgo') or 'RSA'
        
        # Try parsing key size
        key_size = r.get('keySize') or r.get('key_size') or raw.get('keySize') or raw.get('key_size') or 2048
        try:
            key_size = int(key_size)
        except:
            key_size = 2048
            
        pqc = r.get('pqcDetected') or r.get('pqc_detected') or raw.get('pqcDetected') or raw.get('pqc_detected') or False
        
        score = 0
        if pqc:
            score += 0
        else:
            if algo == 'RSA':
                if key_size <= 1024:
                    score += 85
                elif key_size <= 2048:
                    score += 75
                elif key_size <= 3072:
                    score += 60
                else:
                    score += 45
            elif algo in ['ECDSA', 'ECC']:
                score += 70 if key_size <= 256 else 55
            elif algo == 'DSA':
                score += 85
            else:
                score += 60

        if tls == 'TLS 1.0':
            score += 20
        elif tls == 'TLS 1.1':
            score += 15
        elif tls == 'TLS 1.2':
            score += 5

        if 'CBC' in cipher:
            score += 15
        if 'RC4' in cipher:
            score += 25
        if 'DES' in cipher and '3DES' not in cipher:
            score += 20
        if '3DES' in cipher:
            score += 20
        if 'NULL' in cipher or 'EXPORT' in cipher:
            score += 30

        if '128' in cipher and '256' not in cipher:
            score += 10

        if 'CHACHA20' in cipher:
            score -= 10
        if 'GCM' in cipher and 'CBC' not in cipher:
            score -= 5
        if tls == 'TLS 1.3':
            score -= 8
        if pqc:
            score -= 30

        score = max(0, min(100, score))
        scores.append(score)
        
        if score >= 51:
            cbom_count += 1
        elif score <= 25:
            low_count += 1
    avg_score = sum(scores) / len(scores) if scores else 0

    # Calculate NIST PQC compliance score on the backend to match the frontend rating
    nist_score = 0
    if total > 0:
        all_tls13 = all(r.get('tls') == 'TLS 1.3' for r in scan_results)
        any_tls13 = any(r.get('tls') == 'TLS 1.3' for r in scan_results)
        nist_score += 1.0 if all_tls13 else (0.5 if any_tls13 else 0.0)

        all_aead = all(r.get('cipher') and any(x in r.get('cipher').upper() for x in ['GCM', 'CHACHA']) for r in scan_results)
        any_aead = any(r.get('cipher') and any(x in r.get('cipher').upper() for x in ['GCM', 'CHACHA']) for r in scan_results)
        nist_score += 1.0 if all_aead else (0.5 if any_aead else 0.0)

        all_pfs = all(r.get('cipher') and any(x in r.get('cipher').upper() for x in ['ECDHE', 'DHE']) or r.get('tls') == 'TLS 1.3' for r in scan_results)
        nist_score += 1.0 if all_pfs else 0.0

        all_key_ok = all((r.get('keySize') or r.get('key_size') or 0) >= 3072 for r in scan_results)
        any_key_ok = any((r.get('keySize') or r.get('key_size') or 0) >= 3072 for r in scan_results)
        nist_score += 1.0 if all_key_ok else (0.5 if any_key_ok else 0.0)

        any_kyber = any(r.get('pqcDetected') or r.get('pqc_detected') or (r.get('keyExchange') and 'kyber' in r.get('keyExchange').lower()) or (r.get('key_exchange') and 'kyber' in r.get('key_exchange').lower()) for r in scan_results)
        nist_score += 1.0 if any_kyber else 0.0

        any_dilith = any(r.get('certAlgo') in ['CRYSTALS-Dilithium', 'ML-DSA'] or r.get('cert_algo') in ['CRYSTALS-Dilithium', 'ML-DSA'] for r in scan_results)
        nist_score += 1.0 if any_dilith else 0.0

        all_cert_ok = all(r.get('daysLeft') is None or r.get('days_left') is None or (r.get('daysLeft') or r.get('days_left') or 0) > 0 for r in scan_results)
        nist_score += 1.0 if all_cert_ok else 0.0

        no_cbc = all(not r.get('cipher') or not any(x in r.get('cipher').upper() for x in ['CBC', 'RC4', '3DES']) for r in scan_results)
        some_cbc = any(r.get('cipher') and any(x in r.get('cipher').upper() for x in ['CBC', 'RC4']) for r in scan_results)
        nist_score += 1.0 if no_cbc else (0.5 if not some_cbc else 0.0)

    ent_score = round(((1 - (avg_score / 100)) * 700) + ((nist_score / 8) * 300))
    posture_pct = round((low_count / total) * 100)
    return {
        "assets": total,
        "rating": f"{ent_score}/1000",
        "posture": f"{posture_pct}%",
        "cbom": cbom_count
    }

def build_pdf_report(report_title, stats_data, scan_results):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    brand_red = colors.HexColor('#9b1c2e')
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=brand_red,
        spaceAfter=8
    )
    
    sub_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#8a7060'),
        spaceAfter=15
    )
    
    h2_style = ParagraphStyle(
        'Heading2_Custom',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=brand_red,
        spaceBefore=10,
        spaceAfter=8
    )
    
    body_style = ParagraphStyle(
        'Body_Custom',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#333333'),
        spaceAfter=4
    )
    
    story.append(Paragraph(report_title, title_style))
    story.append(Paragraph(f"Generated at: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Q-Sentinel Cryptographic Audit Report", sub_style))
    
    stats_table_data = [
        [
            Paragraph("<b>Assets Scanned</b>", body_style),
            Paragraph("<b>Cyber Rating</b>", body_style),
            Paragraph("<b>PQC Posture</b>", body_style),
            Paragraph("<b>CBOM Issues</b>", body_style)
        ],
        [
            Paragraph(f"<font size=12 color='#9b1c2e'><b>{stats_data.get('assets', 0)}</b></font>", body_style),
            Paragraph(f"<font size=12 color='#10b981'><b>{stats_data.get('rating', '--')}</b></font>", body_style),
            Paragraph(f"<font size=12 color='#f59e0b'><b>{stats_data.get('posture', '--')}</b></font>", body_style),
            Paragraph(f"<font size=12 color='#ef4444'><b>{stats_data.get('cbom', 0)}</b></font>", body_style)
        ]
    ]
    
    stats_table = Table(stats_table_data, colWidths=[135, 135, 135, 135])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#faf6f0')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e8b84b')),
        ('BOX', (0,0), (-1,-1), 1, brand_red),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    
    story.append(Paragraph("Executive Summary", h2_style))
    story.append(stats_table)
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("Scanned Cryptographic Components", h2_style))
    
    headers = ["Asset", "TLS", "Cipher Suite", "Key Size", "CA", "Expiry", "PQC Risk", "Action"]
    table_rows = [[Paragraph(f"<b>{h}</b>", ParagraphStyle('H', parent=body_style, textColor=colors.white)) for h in headers]]
    
    for r in scan_results:
        raw = r.get('raw_response') or {}
        
        host_val = r.get('host') or r.get('hostname') or raw.get('host') or raw.get('hostname') or 'N/A'
        tls_val = r.get('tls') or r.get('tls_version') or raw.get('tls') or raw.get('tls_version') or 'N/A'
        cipher_val = r.get('cipher') or r.get('cipher_suite') or raw.get('cipher') or raw.get('cipher_suite') or 'N/A'
        key_size_val = r.get('keySize') or r.get('key_size') or raw.get('keySize') or raw.get('key_size') or 'N/A'
        issuer_val = r.get('issuer') or raw.get('issuer') or 'N/A'
        expiry_val = r.get('expiry') or raw.get('expiry') or 'N/A'
        
        risk = str(r.get('quantumRisk') or r.get('quantum_risk') or raw.get('quantumRisk') or raw.get('quantum_risk') or 'UNKNOWN').upper()
        
        risk_color = '#10b981'
        if risk in ['CRITICAL', 'HIGH']:
            risk_color = '#ef4444'
        elif risk == 'MODERATE':
            risk_color = '#f59e0b'
            
        action = r.get('action') or r.get('mitigation') or raw.get('action') or raw.get('mitigation') or "Migrate RSA -> CRYSTALS-Kyber-1024"
        
        row = [
            Paragraph(host_val, ParagraphStyle('AS', parent=body_style, fontSize=7)),
            Paragraph(tls_val, body_style),
            Paragraph(cipher_val, ParagraphStyle('C', parent=body_style, fontSize=6)),
            Paragraph(str(key_size_val), body_style),
            Paragraph(issuer_val, ParagraphStyle('I', parent=body_style, fontSize=6)),
            Paragraph(expiry_val, body_style),
            Paragraph(f"<font color='{risk_color}'><b>{risk}</b></font>", body_style),
            Paragraph(action, ParagraphStyle('A', parent=body_style, fontSize=6)),
        ]
        table_rows.append(row)
        
    col_widths = [70, 35, 110, 40, 65, 55, 50, 115]
    result_table = Table(table_rows, colWidths=col_widths, repeatRows=1)
    result_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), brand_red),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e8b84b')),
        ('BOX', (0,0), (-1,-1), 1, brand_red),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    
    story.append(result_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

def build_excel_report(scan_results):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cryptographic Scan"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    headers = ["Asset", "TLS Version", "Cipher Suite", "Key Size", "CA (Issuer)", "Expiry", "Quantum Risk Level", "Recommended Action"]
    ws.append(headers)
    
    brand_red = "9B1C2E"
    header_fill = PatternFill(start_color=brand_red, end_color=brand_red, fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    for r in scan_results:
        raw = r.get('raw_response') or {}
        host_val = r.get('host') or r.get('hostname') or raw.get('host') or raw.get('hostname') or 'N/A'
        tls_val = r.get('tls') or r.get('tls_version') or raw.get('tls') or raw.get('tls_version') or 'N/A'
        cipher_val = r.get('cipher') or r.get('cipher_suite') or raw.get('cipher') or raw.get('cipher_suite') or 'N/A'
        key_size_val = r.get('keySize') or r.get('key_size') or raw.get('keySize') or raw.get('key_size') or 'N/A'
        issuer_val = r.get('issuer') or raw.get('issuer') or 'N/A'
        expiry_val = r.get('expiry') or raw.get('expiry') or 'N/A'
        risk_val = r.get('quantumRisk') or r.get('quantum_risk') or raw.get('quantumRisk') or raw.get('quantum_risk') or 'UNKNOWN'
        action_val = r.get('action') or r.get('mitigation') or raw.get('action') or raw.get('mitigation') or "Migrate RSA -> CRYSTALS-Kyber-1024"
        
        row_data = [host_val, tls_val, cipher_val, key_size_val, issuer_val, expiry_val, risk_val.upper(), action_val]
        ws.append(row_data)
        
        curr_row = ws.max_row
        risk_upper = risk_val.upper()
        risk_fill = None
        risk_font_color = "000000"
        if "CRITICAL" in risk_upper or "HIGH" in risk_upper:
            risk_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
            risk_font_color = "C0392B"
        elif "MODERATE" in risk_upper:
            risk_fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
            risk_font_color = "D35400"
        elif "LOW" in risk_upper or "READY" in risk_upper or "SAFE" in risk_upper:
            risk_fill = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
            risk_font_color = "27AE60"
            
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=curr_row, column=col_num)
            cell.border = thin_border
            cell.alignment = left_align
            if col_num == 7:
                if risk_fill:
                    cell.fill = risk_fill
                cell.font = Font(name="Arial", size=10, bold=True, color=risk_font_color)
                cell.alignment = center_align
            elif col_num in [2, 4, 6]:
                cell.alignment = center_align
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
        
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file.getvalue()

def calculate_initial_next_run(date_str, time_str):
    try:
        clean_time = time_str.split("(")[0].strip()
        dt_str = f"{date_str} {clean_time}"
        dt = datetime.datetime.strptime(dt_str, "%Y-%m-%d %I:%M %p")
        dt = dt - datetime.timedelta(hours=5, minutes=30)
        return dt.replace(tzinfo=datetime.timezone.utc)
    except Exception as e:
        print(f"[SCHEDULE] Date/time parsing error: {e}")
        return datetime.datetime.now(datetime.timezone.utc)

def run_scheduler_loop():
    print("[SCHEDULER] Background scheduler thread started.")
    while True:
        try:
            schedules = db.get_active_schedules()
            now = datetime.datetime.now(datetime.timezone.utc)
            for sched in schedules:
                next_run = sched.get("next_run")
                if next_run:
                    if next_run.tzinfo is None:
                        next_run = next_run.replace(tzinfo=datetime.timezone.utc)
                    
                    if now >= next_run:
                        print(f"[SCHEDULER] Executing schedule: {sched.get('report_type')} for {sched.get('email')}")
                        
                        hosts = db.get_all_hosts()
                        scan_results = []
                        for h in hosts:
                            last_scan = db.get_last_scan_for_host(h['hostname'])
                            if last_scan:
                                scan_results.append(last_scan)
                        
                        stats_data = compute_scan_stats(scan_results)
                        
                        pdf_bytes = build_pdf_report(
                            sched.get('report_type', 'Scheduled Security Report'),
                            stats_data,
                            scan_results
                        )
                        
                        to_email = sched.get('email')
                        subject = f"Scheduled Report: {sched.get('report_type')}"
                        text_content = f"Please find attached your scheduled {sched.get('report_type')} from Q-Sentinel."
                        html_content = f"""
                            <div style="font-family: Arial, sans-serif; max-width: 500px; margin: 0 auto; padding: 20px; border: 1px solid #e0e0e0; border-radius: 8px; background-color: #0d0a0b; color: #f5e6d3;">
                              <h2 style="color: #e8b84b; text-align: center;">Q-Sentinel Security</h2>
                              <hr style="border: none; border-top: 1px solid rgba(232, 184, 75, 0.2); margin: 20px 0;" />
                              <p>Hello Officer,</p>
                              <p>Please find attached your scheduled <b>{sched.get('report_type')}</b> generated automatically by the Q-Sentinel platform.</p>
                              <p style="font-size: 11px; color: #8a7060; text-align: center;">Punjab National Bank Cybersecurity Division</p>
                            </div>
                        """
                        send_email(to_email, subject, text_content, html_content, pdf_bytes, "qsentinel-scheduled-report.pdf")
                        
                        save_path = sched.get('save_path')
                        if save_path:
                            try:
                                clean_dir = save_path.strip().lstrip('/\\')
                                if os.environ.get("VERCEL"):
                                    clean_dir = os.path.join("/tmp", clean_dir)
                                if not os.path.exists(clean_dir):
                                    os.makedirs(clean_dir, exist_ok=True)
                                
                                filename = f"report_{sched.get('report_type').replace(' ', '_')}_{now.strftime('%Y%m%d_%H%M%S')}.pdf"
                                full_file_path = os.path.join(clean_dir, filename)
                                with open(full_file_path, "wb") as f:
                                    f.write(pdf_bytes)
                                print(f"[SCHEDULER] Saved report locally to {full_file_path}")
                            except Exception as file_err:
                                print(f"[SCHEDULER] Error saving report locally on serverless host: {file_err}")
                            
                        freq = sched.get('frequency', 'Weekly').lower()
                        if freq == 'daily':
                            next_run_new = next_run + datetime.timedelta(days=1)
                        elif freq == 'weekly':
                            next_run_new = next_run + datetime.timedelta(days=7)
                        elif freq == 'monthly':
                            next_run_new = next_run + datetime.timedelta(days=30)
                        else:
                            next_run_new = next_run + datetime.timedelta(days=7)
                            
                        sched['next_run'] = next_run_new
                        db.save_schedule(sched)
        except Exception as e:
            print(f"[SCHEDULER ERROR] {e}")
        time.sleep(60)

# Start background scheduler
scheduler_thread = threading.Thread(target=run_scheduler_loop, daemon=True)
scheduler_thread.start()


@app.route("/api/export-pdf", methods=["POST"])
def export_pdf():
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401
    
    data = request.get_json(force=True, silent=True) or {}
    title = data.get("title", "Q-Sentinel Security Report")
    stats = data.get("stats", {})
    scan_results = data.get("scanResults", [])
    
    if not scan_results:
        hosts = db.get_all_hosts()
        for h in hosts:
            last_scan = db.get_last_scan_for_host(h['hostname'])
            if last_scan:
                scan_results.append(last_scan)
                
    stats = compute_scan_stats(scan_results)
        
    try:
        pdf_data = build_pdf_report(title, stats, scan_results)
        from flask import send_file
        return send_file(
            io.BytesIO(pdf_data),
            mimetype="application/pdf",
            as_attachment=True,
            download_name="qsentinel-security-report.pdf"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/schedule-report", methods=["POST"])
def schedule_report():
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401
        
    data = request.get_json(force=True, silent=True) or {}
    report_type = data.get("report_type")
    frequency = data.get("frequency")
    assets = data.get("assets")
    sections = data.get("sections", [])
    date_str = data.get("date")
    time_str = data.get("time")
    email = data.get("email")
    save_path = data.get("save_path")
    enabled = data.get("enabled", True)
    
    if not report_type or not frequency or not email:
        return jsonify({"error": "Missing required fields"}), 400
        
    next_run = calculate_initial_next_run(date_str, time_str)
    
    schedule_doc = {
        "report_type": report_type,
        "frequency": frequency,
        "assets": assets,
        "sections": sections,
        "date": date_str,
        "time": time_str,
        "email": email,
        "save_path": save_path,
        "enabled": enabled,
        "next_run": next_run
    }
    
    try:
        db.save_schedule(schedule_doc)
        return jsonify({"success": True, "message": "Schedule saved successfully!"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/generate-demand-report", methods=["POST"])
def generate_demand_report():
    if not check_api_key():
        return jsonify({"error": "Unauthorized"}), 401
        
    data = request.get_json(force=True, silent=True) or {}
    report_type = data.get("report_type")
    file_format = data.get("format", "PDF")
    include_charts = data.get("include_charts", True)
    password_protect = data.get("password_protect", False)
    email_enabled = data.get("email_enabled", False)
    email = data.get("email", "")
    save_enabled = data.get("save_enabled", False)
    save_path = data.get("save_path", "")
    
    hosts = db.get_all_hosts()
    scan_results = []
    for h in hosts:
        last_scan = db.get_last_scan_for_host(h['hostname'])
        if last_scan:
            scan_results.append(last_scan)
            
    stats_data = compute_scan_stats(scan_results)
    
    filename = f"qsentinel_{report_type.replace(' ', '_').lower()}"
    
    try:
        if file_format == "PDF":
            report_bytes = build_pdf_report(report_type, stats_data, scan_results)
            mimetype = "application/pdf"
            filename += ".pdf"
        elif file_format == "Excel":
            report_bytes = build_excel_report(scan_results)
            mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename += ".xlsx"
        elif file_format == "JSON":
            report_json = {
                "report_type": report_type,
                "stats": stats_data,
                "scan_results": scan_results,
                "generated_at": datetime.datetime.now().isoformat()
            }
            report_bytes = json.dumps(report_json, indent=2).encode('utf-8')
            mimetype = "application/json"
            filename += ".json"
        else:
            csv_lines = ["Asset,TLS Version,Cipher Suite,Key Size,CA,Expiry,Quantum Risk,Action"]
            for r in scan_results:
                raw = r.get('raw_response') or {}
                host_val = r.get('host') or r.get('hostname') or raw.get('host') or raw.get('hostname') or 'N/A'
                tls_val = r.get('tls') or r.get('tls_version') or raw.get('tls') or raw.get('tls_version') or 'N/A'
                cipher_val = r.get('cipher') or r.get('cipher_suite') or raw.get('cipher') or raw.get('cipher_suite') or 'N/A'
                key_size_val = r.get('keySize') or r.get('key_size') or raw.get('keySize') or raw.get('key_size') or 'N/A'
                issuer_val = r.get('issuer') or raw.get('issuer') or 'N/A'
                expiry_val = r.get('expiry') or raw.get('expiry') or 'N/A'
                risk_val = r.get('quantumRisk') or r.get('quantum_risk') or raw.get('quantumRisk') or raw.get('quantum_risk') or 'UNKNOWN'
                action_val = r.get('action') or r.get('mitigation') or raw.get('action') or raw.get('mitigation') or "Migrate RSA -> CRYSTALS-Kyber-1024"
                csv_lines.append(f"{host_val.replace(',', ' ')},{tls_val.replace(',', ' ')},{cipher_val.replace(',', ' ')},{key_size_val},{issuer_val.replace(',', ' ')},{expiry_val.replace(',', ' ')},{risk_val},{action_val.replace(',', ' ')}")
            report_bytes = ("\ufeff" + "\n".join(csv_lines)).encode('utf-8')
            mimetype = "text/csv"
            filename += ".csv"
            
        if email_enabled and email:
            subject = f"On-Demand Report: {report_type}"
            text_content = f"Here is the requested on-demand {report_type} from Q-Sentinel."
            html_content = f"""
                <div style="font-family: Arial, sans-serif; max-width: 500px; margin: 0 auto; padding: 20px; border: 1px solid #e0e0e0; border-radius: 8px; background-color: #0d0a0b; color: #f5e6d3;">
                  <h2 style="color: #e8b84b; text-align: center;">Q-Sentinel Security</h2>
                  <hr style="border: none; border-top: 1px solid rgba(232, 184, 75, 0.2); margin: 20px 0;" />
                  <p>Hello Officer,</p>
                  <p>Please find attached the on-demand <b>{report_type}</b> you generated.</p>
                  <p style="font-size: 11px; color: #8a7060; text-align: center;">Punjab National Bank Cybersecurity Division</p>
                </div>
            """
            send_email(email, subject, text_content, html_content, report_bytes, filename)
            
        if save_enabled and save_path:
            try:
                clean_dir = save_path.strip().lstrip('/\\')
                if os.environ.get("VERCEL"):
                    clean_dir = os.path.join("/tmp", clean_dir)
                if not os.path.exists(clean_dir):
                    os.makedirs(clean_dir, exist_ok=True)
                full_file_path = os.path.join(clean_dir, filename)
                with open(full_file_path, "wb") as f:
                    f.write(report_bytes)
                print(f"[ON-DEMAND] Saved report locally to {full_file_path}")
            except Exception as file_err:
                print(f"[ON-DEMAND] Error saving report locally on serverless host: {file_err}")
            
        from flask import send_file
        return send_file(
            io.BytesIO(report_bytes),
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    print("\n" + "="*55)
    print("  Q-Sentinel v1.0 Backend + Groq AI")
    print("  URL : http://localhost:5004")
    print("="*55 + "\n")
    app.run(host="0.0.0.0", port=5004, debug=False, use_reloader=False)

